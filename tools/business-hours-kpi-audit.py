import argparse
import json
import math
import re
import statistics
import sys
import time
import urllib.parse
import urllib.request
import unicodedata
from datetime import datetime, time as dtime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "supabase.js"
DEFAULT_XLSX = Path(r"C:\Users\guilherme.alamino\Downloads\horarios_comerciais_todos_os_paises.xlsx")


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text



REGION_GROUPS = ["Brasil", "Argentina", "M\u00e9xico", "LATAM", "USA", "ROW"]
MTTS_TARGET_DAYS = {"Brasil": 4, "Argentina": 5, "M\u00e9xico": 6, "LATAM": 6, "USA": 10, "ROW": 10}
MTFC_TARGET_HOURS = {
    "urgente": 1,
    "alta": 2,
    "media": 3,
    "baixa": 5,
    "muito baixa": 6,
}
OFICINA_DEPT_ID = "1128522000008788112"
EXCLUDED_TICKET_NUMBERS = {"220822", "236429", "236430"}
EXCLUDED_CONTACT_DOMAINS = {"unicorndenmart.com", "webpeak.com.br"}
DEFAULT_AGENT_GROUPS = {"Suporte geral", "Especialista", "Sem dono"}


def is_frontend_excluded_ticket(row):
    if str(row.get("ticket_number") or "") in EXCLUDED_TICKET_NUMBERS:
        return True
    email = str(row.get("email") or row.get("contact_email") or "").lower().strip()
    return any(email.endswith("@" + domain) or ("@" + domain) in email for domain in EXCLUDED_CONTACT_DOMAINS)


def is_frontend_excluded_backlog(row):
    return row.get("department_id") == OFICINA_DEPT_ID or is_frontend_excluded_ticket(row)


def is_default_agent_group(row):
    return (row.get("grupo_operacional_agente") or "") in DEFAULT_AGENT_GROUPS


def canonical_region_value(value):
    key = normalize_text(value)
    if key in {"brasil", "brazil"}:
        return "Brasil"
    if key == "argentina":
        return "Argentina"
    if key in {"mexico", "m\u00e9xico"}:
        return "M\u00e9xico"
    if key in {"latam", "america latina", "america latina exceto argentina e mexico", "latin america"}:
        return "LATAM"
    if key in {"usa", "us", "eua", "estados unidos", "united states"}:
        return "USA"
    if key in {"row", "resto do mundo", "rest of world", "africa", "asia", "europa", "india", "oceania"}:
        return "ROW"
    return ""


def ticket_region_group(row):
    return canonical_region_value(row.get("regiao_grupo") or row.get("region") or row.get("regiao") or "")


def ticket_priority_target_hours(row):
    return MTFC_TARGET_HOURS.get(normalize_text(row.get("priority_standard") or row.get("priority")))


def ticket_region_target_days(row):
    region = ticket_region_group(row)
    return MTTS_TARGET_DAYS.get(region)

def read_supabase_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url = re.search(r"SUPABASE_URL\s*=\s*'([^']+)'", text)
    key = re.search(r"SUPABASE_ANON_KEY\s*=\s*'([^']+)'", text)
    if not url or not key:
        raise RuntimeError("SUPABASE_URL/SUPABASE_ANON_KEY nao encontrados em config/supabase.js")
    return url.group(1).rstrip("/"), key.group(1)


def parse_weekdays(value):
    names = {
        "seg": 0, "ter": 1, "qua": 2, "qui": 3, "sex": 4,
        "sab": 5, "dom": 6,
    }
    text = str(value or "").strip()
    if not text:
        return set()
    parts = re.split(r"\s*[-\u2013\u2014]\s*", text)
    if len(parts) == 2:
        start = names.get(normalize_text(parts[0])[:3])
        end = names.get(normalize_text(parts[1])[:3])
        if start is None or end is None:
            return set()
        days = set()
        d = start
        while True:
            days.add(d)
            if d == end:
                break
            d = (d + 1) % 7
        return days
    days = set()
    for piece in re.split(r"[,;/ ]+", text):
        if not piece:
            continue
        day = names.get(normalize_text(piece)[:3])
        if day is not None:
            days.add(day)
    return days


def to_time(value):
    if isinstance(value, dtime):
        return value
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        return None
    return datetime.strptime(text[:5], "%H:%M").time()


def load_schedules(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_name = next((name for name in wb.sheetnames if normalize_text(name) == "paises"), None)
    if not sheet_name:
        raise RuntimeError(f"Aba Paises nao encontrada. Abas: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    idx = {normalize_text(h): i for i, h in enumerate(headers)}
    required = ["pais", "pais (ingles)", "iso", "inicio local", "fim local", "fuso iana", "dias comerciais tipicos", "duracao tipica (h)"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise RuntimeError(f"Cabecalhos ausentes na planilha: {missing}")
    schedules = {}
    aliases = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        country = row[idx["pais"]]
        country_en = row[idx["pais (ingles)"]]
        iso = row[idx["iso"]]
        start = to_time(row[idx["inicio local"]])
        end = to_time(row[idx["fim local"]])
        tz_name = row[idx["fuso iana"]]
        weekdays = parse_weekdays(row[idx["dias comerciais tipicos"]])
        if not country or not start or not end or not tz_name or not weekdays:
            continue
        duration = float(row[idx["duracao tipica (h)"]] or 0)
        item = {
            "country": str(country),
            "country_en": str(country_en or ""),
            "iso": str(iso or ""),
            "tz": str(tz_name),
            "weekdays": sorted(weekdays),
            "start": start,
            "end": end,
            "daily_hours": duration or ((datetime.combine(datetime.today(), end) - datetime.combine(datetime.today(), start)).seconds / 3600),
        }
        key = normalize_text(country)
        schedules[key] = item
        for alias in (country, country_en, iso):
            if alias:
                aliases[normalize_text(alias)] = key

    manual_aliases = {
        "brasil": "brasil",
        "brazil": "brasil",
        "argentina": "argentina",
        "mexico": "mexico",
        "estados unidos": "estados unidos",
        "usa": "estados unidos",
        "eua": "estados unidos",
        "united states": "estados unidos",
        "india": "india",
    }
    for alias, target in manual_aliases.items():
        if target in schedules:
            aliases[normalize_text(alias)] = target
    return schedules, aliases


def resolve_schedule(row, schedules, aliases):
    candidates = [
        row.get("pais"),
        row.get("regiao"),
        row.get("region"),
        row.get("regiao_grupo"),
    ]
    for value in candidates:
        key = aliases.get(normalize_text(value))
        if key and key in schedules:
            return schedules[key]
    return None


def parse_dt(value):
    if not value:
        return None
    text = str(value)
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def business_minutes_between(start_utc, end_utc, schedule):
    if not start_utc or not end_utc or not schedule:
        return None
    if end_utc < start_utc:
        return None
    if end_utc == start_utc:
        return 0.0
    tz = ZoneInfo(schedule["tz"])
    start_local = start_utc.astimezone(tz)
    end_local = end_utc.astimezone(tz)
    total = 0.0
    cur_date = start_local.date()
    while cur_date <= end_local.date():
        if cur_date.weekday() in set(schedule["weekdays"]):
            day_start = datetime.combine(cur_date, schedule["start"], tz)
            day_end = datetime.combine(cur_date, schedule["end"], tz)
            a = max(start_local, day_start)
            b = min(end_local, day_end)
            if b > a:
                total += (b - a).total_seconds() / 60
        cur_date += timedelta(days=1)
    return total


def in_business_hours(moment_utc, schedule):
    if not moment_utc or not schedule:
        return None
    tz = ZoneInfo(schedule["tz"])
    local = moment_utc.astimezone(tz)
    if local.weekday() not in set(schedule["weekdays"]):
        return False
    start = datetime.combine(local.date(), schedule["start"], tz)
    end = datetime.combine(local.date(), schedule["end"], tz)
    return start <= local < end


def rest_get(url, key, table, params):
    query = urllib.parse.urlencode(params, doseq=True, safe="(),.*")
    req = urllib.request.Request(
        f"{url}/rest/v1/{table}?{query}",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as res:
        return json.loads(res.read().decode("utf-8"))


def fetch_paged(url, key, table, select, extra=None, page_size=1000, max_rows=50000):
    rows = []
    offset = 0
    extra = dict(extra or {})
    while offset < max_rows:
        params = {
            "select": select,
            "limit": str(page_size),
            "offset": str(offset),
            **extra,
        }
        page = rest_get(url, key, table, params)
        if not isinstance(page, list):
            raise RuntimeError(f"Resposta inesperada de {table}: {page}")
        rows.extend(page)
        if len(page) < page_size:
            break
        offset += page_size
        time.sleep(0.05)
    return rows


def num(value):
    try:
        if value is None or value == "":
            return math.nan
        out = float(value)
        return out if math.isfinite(out) else math.nan
    except Exception:
        return math.nan


def avg(values):
    vals = [v for v in values if isinstance(v, (int, float)) and math.isfinite(v)]
    return sum(vals) / len(vals) if vals else None


def pct(part, total):
    return (part / total * 100) if total else None


def pctl(values, q):
    vals = sorted(v for v in values if isinstance(v, (int, float)) and math.isfinite(v))
    if not vals:
        return None
    idx = min(len(vals) - 1, int(math.floor(len(vals) * q)))
    return vals[idx]


def is_closed(row):
    status = normalize_text(row.get("status"))
    return bool(row.get("closed_time")) or status in {"fechado", "resolvido", "closed", "resolved"}


def date_in_range(value, start, end):
    dt = parse_dt(value)
    return bool(dt and start <= dt < end)


def current_week(now):
    local = now.astimezone(ZoneInfo("America/Sao_Paulo"))
    sunday = local.date() - timedelta(days=(local.weekday() + 1) % 7)
    start = datetime.combine(sunday, dtime.min, ZoneInfo("America/Sao_Paulo")).astimezone(timezone.utc)
    end = start + timedelta(days=7)
    return start, end


def current_month(now):
    local = now.astimezone(ZoneInfo("America/Sao_Paulo"))
    start_local = datetime(local.year, local.month, 1, tzinfo=ZoneInfo("America/Sao_Paulo"))
    if local.month == 12:
        end_local = datetime(local.year + 1, 1, 1, tzinfo=ZoneInfo("America/Sao_Paulo"))
    else:
        end_local = datetime(local.year, local.month + 1, 1, tzinfo=ZoneInfo("America/Sao_Paulo"))
    return start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc)


def iso_date_brt(dt):
    return dt.astimezone(ZoneInfo("America/Sao_Paulo")).date().isoformat()


def period_label_brt(start, end_exclusive):
    brt = ZoneInfo("America/Sao_Paulo")
    start_date = start.astimezone(brt).date()
    end_date = (end_exclusive - timedelta(seconds=1)).astimezone(brt).date()
    return f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"


def build_week_periods(now, tickets, count=8):
    current_start, _ = current_week(now)
    periods = []
    for i in range(count):
        start = current_start - timedelta(days=7 * i)
        end = start + timedelta(days=7)
        closed_rows = [r for r in tickets if date_in_range(r.get("closed_time"), start, end)]
        periods.append({
            "key": f"{iso_date_brt(start)}__{iso_date_brt(end - timedelta(days=1))}",
            "label": period_label_brt(start, end),
            "start_utc": start.isoformat(),
            "end_utc": end.isoformat(),
            "metrics": compute_period(tickets, start, end),
            "by_region": summarize_by_region(closed_rows),
            "tickets": business_ticket_metrics(closed_rows),
        })
    return periods


def month_start_end_utc(year, month):
    brt = ZoneInfo("America/Sao_Paulo")
    start = datetime(year, month, 1, tzinfo=brt)
    end = datetime(year + 1, 1, 1, tzinfo=brt) if month == 12 else datetime(year, month + 1, 1, tzinfo=brt)
    return start.astimezone(timezone.utc), end.astimezone(timezone.utc)


def build_month_periods(tickets):
    months = set()
    brt = ZoneInfo("America/Sao_Paulo")
    for row in tickets:
        closed = parse_dt(row.get("closed_time"))
        if not closed:
            continue
        local = closed.astimezone(brt)
        months.add((local.year, local.month))
    periods = []
    for year, month in sorted(months, reverse=True):
        start, end = month_start_end_utc(year, month)
        closed_rows = [r for r in tickets if date_in_range(r.get("closed_time"), start, end)]
        periods.append({
            "key": f"{year:04d}-{month:02d}",
            "label": f"{month:02d}/{year:04d}",
            "start_utc": start.isoformat(),
            "end_utc": end.isoformat(),
            "metrics": compute_period(tickets, start, end),
            "by_region": summarize_by_region(closed_rows),
            "tickets": business_ticket_metrics(closed_rows),
        })
    return periods


def enrich_ticket(row, schedules, aliases, now):
    schedule = resolve_schedule(row, schedules, aliases)
    created = parse_dt(row.get("created_time"))
    closed = parse_dt(row.get("closed_time"))
    mtfc = num(row.get("mtfc_horas_bi"))
    first_response = created + timedelta(hours=mtfc) if created and math.isfinite(mtfc) else None
    mtfc_bh = None
    if schedule and first_response:
        minutes = business_minutes_between(created, first_response, schedule)
        mtfc_bh = minutes / 60 if minutes is not None else None
    mtts_bh_days = None
    if schedule and created and closed:
        minutes = business_minutes_between(created, closed, schedule)
        if minutes is not None:
            mtts_bh_days = (minutes / 60) / (schedule.get("daily_hours") or 8)
    aging_bh_days = None
    if schedule and created and not closed:
        minutes = business_minutes_between(created, now, schedule)
        if minutes is not None:
            aging_bh_days = (minutes / 60) / (schedule.get("daily_hours") or 8)
    return {
        **row,
        "_schedule_country": schedule["country"] if schedule else None,
        "_created_dt": created,
        "_closed_dt": closed,
        "_created_in_bh": in_business_hours(created, schedule) if schedule else None,
        "_closed_in_bh": in_business_hours(closed, schedule) if schedule and closed else None,
        "_mtfc_bh_hours": mtfc_bh,
        "_mtts_bh_days": mtts_bh_days,
        "_aging_bh_days": aging_bh_days,
    }


def compute_closed_metrics(rows):
    closed = [r for r in rows if is_closed(r)]
    ticket_metrics = [business_ticket_metric(r) for r in closed]

    orig_eligible = [m for m in ticket_metrics if m["orig"].get("eligible") is True]
    orig_within = [m for m in orig_eligible if m["orig"].get("within") is True]
    bh_eligible = [m for m in ticket_metrics if m["business_hours"].get("eligible") is True]
    bh_within = [m for m in bh_eligible if m["business_hours"].get("within") is True]

    return {
        "closed": len(ticket_metrics),
        "orig": {
            "eligible": len(orig_eligible),
            "within": len(orig_within),
            "sla_pct": pct(len(orig_within), len(orig_eligible)),
            "avg_mtfc_h": avg(m["orig"].get("mtfc_h") for m in ticket_metrics),
            "avg_mtts_d": avg(m["orig"].get("mtts_d") for m in ticket_metrics),
        },
        "business_hours": {
            "eligible": len(bh_eligible),
            "within": len(bh_within),
            "sla_pct": pct(len(bh_within), len(bh_eligible)),
            "avg_mtfc_h": avg(m["business_hours"].get("mtfc_h") for m in ticket_metrics),
            "avg_mtts_business_days": avg(m["business_hours"].get("mtts_business_days") for m in ticket_metrics),
            "unresolved_schedule": sum(1 for m in ticket_metrics if m["business_hours"].get("unresolved_schedule") is True),
            "fallback_original_eligible": sum(1 for m in bh_eligible if m["business_hours"].get("mtfc_fallback_original") or m["business_hours"].get("mtts_fallback_original")),
            "mtfc_note": "aproximado: created_time + mtfc_horas_bi; quando pais/horario nao e resolvido, preserva o valor original do ticket",
        },
    }

def effective_business_mtfc(row):
    value = row.get("_mtfc_bh_hours")
    if isinstance(value, (int, float)) and math.isfinite(value):
        return value, False
    return num(row.get("mtfc_horas_bi")), True


def effective_business_mtts(row):
    value = row.get("_mtts_bh_days")
    if isinstance(value, (int, float)) and math.isfinite(value):
        return value, False
    return num(row.get("mtts_dias_bi")), True



def ticket_product(row):
    return row.get("produtos") or row.get("product") or ""


def ticket_product_line(row):
    product = normalize_text(ticket_product(row))
    if not product:
        return ""
    if any(term in product for term in [
        "digitalizador", "eagle", "raio-x", "raio x", "scanner",
        "sensor intraoral", "tomografo", "panoramico"
    ]):
        return "imaging"
    if any(term in product for term in [
        "autoclave", "bomba vacuo", "compressor", "consultorio",
        "fotopolimerizador", "micro motor", "pecas de mao", "perfilaxia", "profilaxia"
    ]):
        return "dental"
    return ""

def business_ticket_metric(row):
    mtfc, mtfc_fallback = effective_business_mtfc(row)
    mtts, mtts_fallback = effective_business_mtts(row)
    mtfc_target = num(row.get("meta_mtfc_horas"))
    if not math.isfinite(mtfc_target):
        mtfc_target = ticket_priority_target_hours(row)
    mtts_target = num(row.get("meta_mtts_dias"))
    if not math.isfinite(mtts_target):
        mtts_target = ticket_region_target_days(row)
    eligible = all(isinstance(v, (int, float)) and math.isfinite(v) for v in [mtfc, mtts, mtfc_target, mtts_target])
    within = bool(eligible and mtfc <= mtfc_target and mtts <= mtts_target)
    return {
        "ticket_id": row.get("ticket_id"),
        "ticket_number": row.get("ticket_number"),
        "region": ticket_region_group(row),
        "priority": row.get("priority_standard") or row.get("priority"),
        "department_name": row.get("department_name"),
        "type": row.get("tipo_atendimento"),
        "product": ticket_product(row),
        "product_line": ticket_product_line(row),
        "status": row.get("status"),
        "agent_group": row.get("grupo_operacional_agente"),
        "created_time": row.get("created_time"),
        "closed_time": row.get("closed_time"),
        "orig": {
            "eligible": row.get("is_sla_eligible") is True or row.get("sla_status_bi") in ("Dentro SLA", "Fora SLA"),
            "within": row.get("sla_status_bi") == "Dentro SLA",
            "sla_status": row.get("sla_status_bi"),
            "mtfc_h": num(row.get("mtfc_horas_bi")),
            "mtts_d": num(row.get("mtts_dias_bi")),
        },
        "business_hours": {
            "eligible": eligible,
            "within": within,
            "sla_status": "Dentro SLA" if within else ("Fora SLA" if eligible else "Nao elegivel"),
            "mtfc_h": mtfc,
            "mtts_business_days": mtts,
            "mtfc_fallback_original": mtfc_fallback,
            "mtts_fallback_original": mtts_fallback,
            "mtfc_target_h": mtfc_target,
            "mtts_target_business_days": mtts_target,
            "unresolved_schedule": not bool(row.get("_schedule_country")),
        },
    }


def business_ticket_metrics(rows):
    return [business_ticket_metric(r) for r in rows if is_closed(r)]


def ticket_metric_summary(ticket_metrics):
    rows = list(ticket_metrics or [])
    orig_eligible = [m for m in rows if m.get("orig", {}).get("eligible") is True]
    orig_within = [m for m in orig_eligible if m.get("orig", {}).get("within") is True]
    bh_eligible = [m for m in rows if m.get("business_hours", {}).get("eligible") is True]
    bh_within = [m for m in bh_eligible if m.get("business_hours", {}).get("within") is True]
    return {
        "closed": len(rows),
        "orig": {
            "eligible": len(orig_eligible),
            "within": len(orig_within),
            "sla_pct": pct(len(orig_within), len(orig_eligible)),
            "avg_mtfc_h": avg(m.get("orig", {}).get("mtfc_h") for m in rows),
            "avg_mtts_d": avg(m.get("orig", {}).get("mtts_d") for m in rows),
        },
        "business_hours": {
            "eligible": len(bh_eligible),
            "within": len(bh_within),
            "sla_pct": pct(len(bh_within), len(bh_eligible)),
            "avg_mtfc_h": avg(m.get("business_hours", {}).get("mtfc_h") for m in rows),
            "avg_mtts_business_days": avg(m.get("business_hours", {}).get("mtts_business_days") for m in rows),
            "unresolved_schedule": sum(1 for m in rows if m.get("business_hours", {}).get("unresolved_schedule") is True),
            "fallback_original_eligible": sum(1 for m in bh_eligible if m.get("business_hours", {}).get("mtfc_fallback_original") or m.get("business_hours", {}).get("mtts_fallback_original")),
            "mtfc_note": "aproximado: created_time + mtfc_horas_bi; quando pais/horario nao e resolvido, preserva o valor original do ticket",
        },
    }


def sync_period_metrics_from_tickets(report):
    for period_key, top_key, region_key in [
        ("weekly_periods", "weekly", "weekly_by_region"),
        ("monthly_periods", "monthly", "monthly_by_region"),
    ]:
        periods = report.get(period_key) or []
        for period in periods:
            summary = ticket_metric_summary(period.get("tickets") or [])
            metrics = period.setdefault("metrics", {})
            for key, value in summary.items():
                metrics[key] = value
            by_region = {}
            for region in REGION_GROUPS:
                region_rows = [m for m in period.get("tickets") or [] if m.get("region") == region]
                if region_rows:
                    by_region[region] = ticket_metric_summary(region_rows)
            period["by_region"] = by_region
        if periods:
            report[top_key] = periods[0]["metrics"]
            report[region_key] = periods[0].get("by_region", {})
    return report

def summarize_by_region(rows):
    return {
        region: compute_closed_metrics([r for r in rows if ticket_region_group(r) == region])
        for region in REGION_GROUPS
        if any(ticket_region_group(r) == region for r in rows)
    }


def compute_daily(backlog_rows):
    total = len(backlog_rows)
    no_response = sum(1 for r in backlog_rows if not math.isfinite(num(r.get("mtfc_horas_bi"))))
    over10 = sum(1 for r in backlog_rows if num(r.get("aging_backlog_dias")) > 10)
    p12 = sum(1 for r in backlog_rows if normalize_text(r.get("priority_standard") or r.get("priority")) in {"urgente", "alta", "p1", "p2", "urgent", "high"})
    return {
        "total_backlog": total,
        "sem_primeira_resposta": no_response,
        "p1_p2_abertos": p12,
        "orig": {
            "aging_medio_dias_corridos": avg(num(r.get("aging_backlog_dias")) for r in backlog_rows),
            "maior_10_dias_corridos": over10,
        },
        "business_hours": {
            "aging_medio_dias_uteis": avg(r.get("_aging_bh_days") for r in backlog_rows),
            "maior_10_dias_uteis": sum(1 for r in backlog_rows if (r.get("_aging_bh_days") or -1) > 10),
            "unresolved_schedule": sum(1 for r in backlog_rows if not r.get("_schedule_country")),
        },
    }


def compute_period(tickets, start, end):
    created = [r for r in tickets if date_in_range(r.get("created_time"), start, end)]
    closed = [r for r in tickets if date_in_range(r.get("closed_time"), start, end)]
    metrics = compute_closed_metrics(closed)
    metrics["created"] = len(created)
    metrics["created_inside_business_hours"] = sum(1 for r in created if r.get("_created_in_bh") is True)
    metrics["created_unresolved_schedule"] = sum(1 for r in created if not r.get("_schedule_country"))
    metrics["closed_inside_business_hours"] = sum(1 for r in closed if r.get("_closed_in_bh") is True)
    metrics["closed_unresolved_schedule"] = sum(1 for r in closed if not r.get("_schedule_country"))
    return metrics


def compute_kanban(intervals, schedules, aliases, now):
    enriched = []
    for row in intervals:
        schedule = resolve_schedule(row, schedules, aliases)
        entered = parse_dt(row.get("status_entered_at"))
        exited = parse_dt(row.get("status_exited_at")) or now
        bh_hours = None
        if schedule and entered and exited:
            minutes = business_minutes_between(entered, exited, schedule)
            bh_hours = minutes / 60 if minutes is not None else None
        enriched.append({**row, "_schedule_country": schedule["country"] if schedule else None, "_bh_hours": bh_hours})
    statuses = sorted(set(r.get("status") or "" for r in enriched if r.get("status")))
    out = {}
    for status in statuses:
        rows = [r for r in enriched if r.get("status") == status]
        orig = [num(r.get("status_duration_hours")) for r in rows]
        bh = [r.get("_bh_hours") for r in rows]
        out[status] = {
            "intervalos": len(rows),
            "tickets_distintos": len(set(str(r.get("ticket_id")) for r in rows if r.get("ticket_id") is not None)),
            "orig_avg_h": avg(orig),
            "orig_median_h": statistics.median([v for v in orig if math.isfinite(v)]) if any(math.isfinite(v) for v in orig) else None,
            "orig_p90_h": pctl(orig, 0.9),
            "bh_avg_h": avg(bh),
            "bh_median_h": statistics.median([v for v in bh if isinstance(v, (int, float)) and math.isfinite(v)]) if any(isinstance(v, (int, float)) and math.isfinite(v) for v in bh) else None,
            "bh_p90_h": pctl(bh, 0.9),
            "unresolved_schedule": sum(1 for r in rows if not r.get("_schedule_country")),
        }
    return out


def round_deep(value):
    if isinstance(value, float):
        return round(value, 4) if math.isfinite(value) else None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: round_deep(v) for k, v in value.items()}
    if isinstance(value, list):
        return [round_deep(v) for v in value]
    return value


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output", default="")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise RuntimeError(f"Planilha nao encontrada: {xlsx_path}")

    schedules, aliases = load_schedules(xlsx_path)
    url, key = read_supabase_config()
    now = datetime.now(timezone.utc)
    week_start, week_end = current_week(now)
    month_start, month_end = current_month(now)

    ticket_select = ",".join([
        "ticket_id", "ticket_number", "status", "priority", "priority_standard",
        "department_name", "contact_email", "pais", "region", "regiao", "regiao_grupo",
        "created_time", "closed_time", "due_date", "response_due_date",
        "mtfc_horas_bi", "mtts_dias_bi", "aging_backlog_dias",
        "is_open", "is_closed", "is_sla_eligible", "meta_mtts_dias",
        "meta_mtfc_horas", "sla_status_bi", "grupo_operacional_agente",
        "produtos", "tipo_atendimento", "categoria",
    ])
    tickets_raw = fetch_paged(
        url,
        key,
        "vw_tickets_bi_base",
        ticket_select,
        {"or": "(is_weekly_report_filter_included.eq.true,department_name.eq.SAC)", "order": "created_time.desc"},
    )
    backlog_raw = fetch_paged(
        url,
        key,
        "vw_dashboard_bi_backlog",
        "*",
        {"order": "aging_backlog_dias.desc"},
    )
    interval_select = ",".join([
        "ticket_id", "ticket_number", "status", "pais", "region", "regiao",
        "department_name", "priority", "created_time", "closed_time",
        "status_entered_at", "status_exited_at", "status_duration_hours",
        "is_deleted",
    ])
    intervals_raw = fetch_paged(
        url,
        key,
        "vw_ticket_status_intervals",
        interval_select,
        {"is_deleted": "eq.false", "status_duration_hours": "gte.0"},
    )

    tickets = [enrich_ticket(r, schedules, aliases, now) for r in tickets_raw if not is_frontend_excluded_ticket(r) and is_default_agent_group(r)]
    backlog = [enrich_ticket(r, schedules, aliases, now) for r in backlog_raw if not is_frontend_excluded_backlog(r) and is_default_agent_group(r)]
    weekly_periods = build_week_periods(now, tickets, 8)
    monthly_periods = build_month_periods(tickets)
    current_week_metrics = weekly_periods[0]["metrics"] if weekly_periods else compute_period(tickets, week_start, week_end)
    current_month_metrics = monthly_periods[0]["metrics"] if monthly_periods else compute_period(tickets, month_start, month_end)
    current_week_by_region = weekly_periods[0]["by_region"] if weekly_periods else summarize_by_region([r for r in tickets if date_in_range(r.get("closed_time"), week_start, week_end)])
    current_month_by_region = monthly_periods[0]["by_region"] if monthly_periods else summarize_by_region([r for r in tickets if date_in_range(r.get("closed_time"), month_start, month_end)])

    report = {
        "generated_at": now.isoformat(),
        "source": {
            "supabase_url": url,
            "tickets_rows": len(tickets),
            "backlog_rows": len(backlog),
            "kanban_intervals_rows": len(intervals_raw),
            "business_hours_workbook": str(xlsx_path),
        },
        "rules": {
            "mtts_business_days": "business_hours_between(created_time, closed_time) / local_daily_hours",
            "mtfc_business_hours": "business_hours_between(created_time, created_time + mtfc_horas_bi); aproximado por falta de first_response_at na view",
            "event_inside_business_hours": "created_time/closed_time dentro dos dias e janela locais do pais",
            "unresolved_country": "tickets com pais/regiao agregados sem pais especifico nao entram nos tempos BH",
        },
        "periods": {
            "current_week": {"start_utc": week_start.isoformat(), "end_utc": week_end.isoformat()},
            "current_month": {"start_utc": month_start.isoformat(), "end_utc": month_end.isoformat()},
        },
        "daily": compute_daily(backlog),
        "weekly": current_week_metrics,
        "weekly_by_region": current_week_by_region,
        "weekly_periods": weekly_periods,
        "monthly": current_month_metrics,
        "monthly_by_region": current_month_by_region,
        "monthly_periods": monthly_periods,
        "kanban": compute_kanban(intervals_raw, schedules, aliases, now),
        "country_resolution": {
            "tickets_unresolved_schedule": sum(1 for r in tickets if not r.get("_schedule_country")),
            "backlog_unresolved_schedule": sum(1 for r in backlog if not r.get("_schedule_country")),
        },
    }

    report = round_deep(report)
    report = sync_period_metrics_from_tickets(report)
    if args.output:
        out = Path(args.output)
        latest_out = None
    else:
        out_dir = ROOT / "reports"
        out_dir.mkdir(exist_ok=True)
        stamp = now.astimezone(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d-%H%M%S")
        out = out_dir / f"business-hours-kpi-audit-{stamp}.json"
        latest_out = ROOT / "01_KPI" / "KPI_V2" / "data" / "business-hours-kpi-audit-latest.json"
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    out.write_text(payload, encoding="utf-8")
    if latest_out:
        latest_out.parent.mkdir(parents=True, exist_ok=True)
        latest_out.write_text(payload, encoding="utf-8")
    print(json.dumps({
        "output": str(out),
        "latest_output": str(latest_out) if latest_out else "",
        "source": report["source"],
        "periods": report["periods"],
        "daily": report["daily"],
        "weekly": report["weekly"],
        "monthly": report["monthly"],
        "country_resolution": report["country_resolution"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        sys.exit(1)
